"""
file_reader.py — Data File Text Extractor
==========================================
Extracts readable text from uploaded data files and returns a
formatted context block for injection into the Council debate.

Supported formats
-----------------
  CSV / TSV          stdlib csv   (no extra dep)
  TXT / MD / LOG     raw utf-8    (no extra dep)
  JSON / JSONL       stdlib json  (no extra dep)
  PDF                pypdf        (pip install pypdf)
  Excel .xlsx/.xls   openpyxl     (pip install openpyxl)
  DOCX               python-docx  (pip install python-docx)

All optional-dependency parsers degrade gracefully: if the library
is not installed the file is returned as an "unsupported" warning
instead of crashing the app.
"""

import csv
import io
import json
from typing import List, Tuple

from glossary import (
    FILE_UPLOAD_CONTEXT_HEADER,
    FILE_UPLOAD_MAX_ROWS_CSV,
    FILE_UPLOAD_MAX_TOTAL_CHARS,
)

# ---------------------------------------------------------------------------
# Per-format parsers — each returns (text: str, warning: str | None)
# ---------------------------------------------------------------------------

def _parse_text(raw: bytes) -> Tuple[str, str | None]:
    return raw.decode("utf-8", errors="replace"), None


def _parse_json(raw: bytes) -> Tuple[str, str | None]:
    try:
        obj = json.loads(raw.decode("utf-8", errors="replace"))
        text = json.dumps(obj, ensure_ascii=False, indent=2)
        if len(text) > FILE_UPLOAD_MAX_TOTAL_CHARS:
            text = text[:FILE_UPLOAD_MAX_TOTAL_CHARS] + "\n… [נקטע]"
        return text, None
    except json.JSONDecodeError as exc:
        # Fall back to raw text if JSON is malformed
        return raw.decode("utf-8", errors="replace"), f"JSON לא תקני: {exc}"


def _parse_jsonl(raw: bytes) -> Tuple[str, str | None]:
    lines = raw.decode("utf-8", errors="replace").splitlines()
    parsed = []
    for i, line in enumerate(lines[:FILE_UPLOAD_MAX_ROWS_CSV]):
        line = line.strip()
        if not line:
            continue
        try:
            parsed.append(json.dumps(json.loads(line), ensure_ascii=False))
        except json.JSONDecodeError:
            parsed.append(line)
    text = "\n".join(parsed)
    warn = f"מציג {FILE_UPLOAD_MAX_ROWS_CSV} שורות ראשונות מתוך {len(lines)}" if len(lines) > FILE_UPLOAD_MAX_ROWS_CSV else None
    return text, warn


def _parse_csv(raw: bytes, delimiter: str = ",") -> Tuple[str, str | None]:
    text_io = io.StringIO(raw.decode("utf-8", errors="replace"))
    reader  = csv.reader(text_io, delimiter=delimiter)
    rows    = list(reader)
    if not rows:
        return "(קובץ ריק)", None

    headers   = rows[0]
    data_rows = rows[1:]
    total     = len(data_rows)

    # Format as a markdown table (capped at MAX_ROWS_CSV rows)
    preview = data_rows[:FILE_UPLOAD_MAX_ROWS_CSV]
    col_w   = [max(len(str(h)), max((len(str(r[i])) if i < len(r) else 0 for r in preview), default=0))
               for i, h in enumerate(headers)]

    def _row_str(r: list) -> str:
        cells = [str(r[i]) if i < len(r) else "" for i in range(len(headers))]
        return "| " + " | ".join(c.ljust(col_w[i]) for i, c in enumerate(cells)) + " |"

    sep  = "| " + " | ".join("-" * w for w in col_w) + " |"
    lines = [_row_str(headers), sep] + [_row_str(r) for r in preview]
    text  = f"**עמודות:** {len(headers)}  |  **שורות סה״כ:** {total:,}\n\n" + "\n".join(lines)
    warn  = f"מציג {len(preview):,} שורות ראשונות מתוך {total:,}" if total > FILE_UPLOAD_MAX_ROWS_CSV else None
    return text, warn


def _parse_pdf(raw: bytes) -> Tuple[str, str | None]:
    try:
        import pypdf  # type: ignore
    except ImportError:
        return "", "pypdf לא מותקן — הרץ: pip install pypdf"

    reader = pypdf.PdfReader(io.BytesIO(raw))
    pages_text = []
    for i, page in enumerate(reader.pages):
        extracted = page.extract_text() or ""
        pages_text.append(f"--- עמוד {i + 1} ---\n{extracted}")

    text = "\n\n".join(pages_text)
    if len(text) > FILE_UPLOAD_MAX_TOTAL_CHARS:
        text = text[:FILE_UPLOAD_MAX_TOTAL_CHARS] + "\n… [נקטע — הקובץ ארוך מהמגבלה]"
        return text, "ה-PDF נקטע כדי לעמוד במגבלת התווים"
    return text, None


def _parse_excel(raw: bytes, filename: str) -> Tuple[str, str | None]:
    try:
        import openpyxl  # type: ignore
    except ImportError:
        return "", "openpyxl לא מותקן — הרץ: pip install openpyxl"

    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    sheet_blocks = []
    for sheet_name in wb.sheetnames:
        ws   = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        total = len(rows) - 1  # minus header
        preview = rows[:FILE_UPLOAD_MAX_ROWS_CSV + 1]  # header + data
        csv_lines = []
        for row in preview:
            csv_lines.append(", ".join("" if v is None else str(v) for v in row))
        block = f"**גיליון: {sheet_name}**  |  שורות: {total:,}\n\n" + "\n".join(csv_lines)
        if total > FILE_UPLOAD_MAX_ROWS_CSV:
            block += f"\n… [מציג {FILE_UPLOAD_MAX_ROWS_CSV:,} שורות ראשונות]"
        sheet_blocks.append(block)
    wb.close()
    return "\n\n---\n\n".join(sheet_blocks), None


def _parse_docx(raw: bytes) -> Tuple[str, str | None]:
    try:
        import docx  # python-docx  # type: ignore
    except ImportError:
        return "", "python-docx לא מותקן — הרץ: pip install python-docx"

    doc  = docx.Document(io.BytesIO(raw))
    text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    if len(text) > FILE_UPLOAD_MAX_TOTAL_CHARS:
        text = text[:FILE_UPLOAD_MAX_TOTAL_CHARS] + "\n… [נקטע]"
    return text, None


# ---------------------------------------------------------------------------
# Public interface
# ---------------------------------------------------------------------------

def extract_file(filename: str, raw: bytes) -> Tuple[str, str | None]:
    """
    Extract text from a single uploaded file.

    Parameters
    ----------
    filename : str
        Original filename (used to detect format).
    raw : bytes
        Raw file bytes from st.file_uploader.

    Returns
    -------
    Tuple[str, str | None]
        (text_content, warning_or_None)
    """
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    if ext in ("txt", "md", "log", "rst"):
        return _parse_text(raw)
    if ext == "csv":
        return _parse_csv(raw, delimiter=",")
    if ext == "tsv":
        return _parse_csv(raw, delimiter="\t")
    if ext == "json":
        return _parse_json(raw)
    if ext == "jsonl":
        return _parse_jsonl(raw)
    if ext == "pdf":
        return _parse_pdf(raw)
    if ext in ("xlsx", "xls"):
        return _parse_excel(raw, filename)
    if ext == "docx":
        return _parse_docx(raw)

    return "", f"סוג קובץ לא נתמך: .{ext}"


def build_file_context(files: List[Tuple[str, bytes]]) -> Tuple[str, List[str]]:
    """
    Process multiple uploaded files and return a combined context block.

    Parameters
    ----------
    files : list of (filename, raw_bytes)

    Returns
    -------
    Tuple[str, List[str]]
        (context_block, warnings)
    """
    if not files:
        return "", []

    blocks: List[str] = []
    warnings: List[str] = []
    total_chars = 0

    for filename, raw in files:
        text, warn = extract_file(filename, raw)
        if warn:
            warnings.append(f"{filename}: {warn}")
        if not text:
            continue

        remaining = FILE_UPLOAD_MAX_TOTAL_CHARS - total_chars
        if remaining <= 0:
            warnings.append(f"דולג: {filename} — תקציב התווים הכולל מוצה.")
            continue
        if len(text) > remaining:
            text = text[:remaining] + "\n… [נקטע]"
            warnings.append(f"קובץ נקטע: {filename}")

        size_kb = len(raw) // 1024
        blocks.append(f"### קובץ: {filename}  ({size_kb} KB)\n\n{text}")
        total_chars += len(text)

    if not blocks:
        return "", warnings

    header = FILE_UPLOAD_CONTEXT_HEADER.format(
        file_count=len(blocks),
        total_chars=total_chars,
    )
    return header + "\n\n" + "\n\n---\n\n".join(blocks), warnings
